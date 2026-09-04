"""Gera XLSX STAR Ondas de Calor estruturado exatamente pelos itens do Anexo 1."""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from sisclima.core.db import read_table

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "output" / "star" / "STAR_Anexo1_Ondas_de_Calor_Materiais.xlsx"
CSV_MUN = ROOT / "data" / "output" / "star" / "STAR_ondas_calor_municipal_SE35_2026.csv"
JSON_RESUMO = ROOT / "data" / "output" / "star" / "STAR_resumo_indicadores.json"

HEADER_FILL = PatternFill("solid", fgColor="1D357F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1D357F", name="Calibri")
SECTION_FONT = Font(bold=True, size=12, color="1D357F", name="Calibri")
ALT = PatternFill("solid", fgColor="F7F9FC")
STATUS = {
    "Disponível": PatternFill("solid", fgColor="C6EFCE"),
    "Parcial": PatternFill("solid", fgColor="FFEB9C"),
    "Indisponível": PatternFill("solid", fgColor="FFC7CE"),
    "Lacuna": PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Border(
    left=Side(style="thin", color="C8D2E6"),
    right=Side(style="thin", color="C8D2E6"),
    top=Side(style="thin", color="C8D2E6"),
    bottom=Side(style="thin", color="C8D2E6"),
)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 30


def _autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:40]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_df(ws, df: pd.DataFrame, freeze="A2"):
    if df is None or df.empty:
        ws["A1"] = "Sem dados disponíveis para este item."
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, None if pd.isna(val) else val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if r_idx == 1:
                continue
            if r_idx % 2 == 0:
                cell.fill = ALT
    _style_header(ws, 1)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze
    _autosize(ws)


def _write_kv_table(ws, title: str, rows: list[tuple], start_row: int = 1):
    ws.cell(start_row, 1, title).font = TITLE_FONT
    headers = rows[0]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row + 2, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    for r_idx, row in enumerate(rows[1:], start_row + 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c_idx == 2 and str(val) in STATUS:
                cell.fill = STATUS[str(val)]
            elif r_idx % 2 == 0:
                cell.fill = ALT
    _autosize(ws, 70)
    return start_row + 2 + len(rows)


def main() -> None:
    resumo = json.loads(JSON_RESUMO.read_text(encoding="utf-8")) if JSON_RESUMO.exists() else {}
    mun = pd.read_csv(CSV_MUN) if CSV_MUN.exists() else pd.DataFrame()

    wb = Workbook()

    # -------- Capa / Índice --------
    ws0 = wb.active
    ws0.title = "00_Capa_Indice"
    ws0["A1"] = "Oficina STAR — MATERIAIS NECESSÁRIOS"
    ws0["A1"].font = TITLE_FONT
    ws0["A2"] = "Tema: Ondas de Calor"
    ws0["A2"].font = SECTION_FONT
    ws0["A3"] = "Estado: Mato Grosso | CIEVS-MT / SES-MT"
    ws0["A4"] = "Referência: SE 35/2026 (03/09/2026) | Série sanitária disponível a partir de 2024 quando indicado"
    ws0["A5"] = "Estrutura do arquivo: cada aba corresponde a um bloco/item solicitado no Anexo 1."
    ws0["A7"] = "Aba"
    ws0["B7"] = "Conteúdo (item do Anexo)"
    _style_header(ws0, 7)
    indice = [
        ("01_A_Analise_Epidemiologica", "Bloco A — Análise epidemiológica e climática (formulário item a item)"),
        ("02_A_Limiares", "Definição dos limiares (duração, intensidade, áreas)"),
        ("03_A_Municipal_Exposicao", "Mapas/dados de exposição, densidade e vulnerabilidade social (municipal)"),
        ("04_A_Tendencias", "Tendências, sazonalidade, períodos de risco e previsão"),
        ("05_A_Saude_Calor", "Atendimentos/internações/óbitos possivelmente relacionados ao calor"),
        ("06_A_Distribuicao_Mun", "Distribuição por município/região de saúde e grupos vulneráveis disponíveis"),
        ("07_A_Sobrecarga", "Demanda/sobrecarga dos serviços de saúde"),
        ("08_A_Outras_Analises", "Outras análises pertinentes"),
        ("09_B_Caracterizacao", "Bloco B — Caracterização do evento e impactos à saúde"),
        ("10_B_Grupos_Vulneraveis", "Grupos com maior vulnerabilidade"),
        ("11_C_Vigilancia_Resposta", "Bloco C — Sistema de vigilância e capacidade de resposta"),
        ("12_Alertas_INMET", "Alertas meteorológicos recentes (INMET)"),
        ("13_Serie_Saude_Mensal", "Série mensal saúde–calor (quando disponível)"),
    ]
    for i, (aba, desc) in enumerate(indice, 8):
        ws0.cell(i, 1, aba).border = THIN
        ws0.cell(i, 2, desc).border = THIN
    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 90

    # -------- Bloco A formulário --------
    wsA = wb.create_sheet("01_A_Analise_Epidemiologica")
    rows_a = [
        ("Item do Anexo", "Status", "Dados / resposta", "Lacuna"),
        (
            "Série histórica de ondas de calor, Tmáx, anomalias e alertas por ano/mês/SE (mín. 5 anos)",
            "Parcial",
            f"Rodada SE 35/2026: {resumo.get('vr_atual')}/142 ({100*resumo.get('vr_atual',0)/142:.1f}%) vermelho/roxo; "
            f"projeção ~7d {resumo.get('vr_proj')}/142; Tmáx máx {resumo.get('tmax_max')} °C; "
            f"hist_clima local {resumo.get('hist_clima_periodo')}; alertas INMET recentes n={resumo.get('inmet_n', 'ND')}.",
            "Sem série climatológica municipal ≥5 anos nem arquivo histórico completo de alertas/anomalias oficiais.",
        ),
        (
            "Definição dos limiares (duração, intensidade, áreas abrangidas)",
            "Disponível",
            "Ver aba 02_A_Limiares. Onda: ≥2 dias consecutivos acima do P95; UTCI/Tmáx/risco cumulativo 3d/EHF.",
            "P95 pode ser proxy operacional na ausência de climatologia longa.",
        ),
        (
            "Mapas temáticos: exposição, ilhas de calor, arborização, densidade, vulnerabilidade social",
            "Parcial",
            "Disponível exposição municipal, densidade, idosos, ruralidade e índice de vulnerabilidade (aba 03). Mapas do boletim SE 35.",
            "Ilhas de calor urbanas e baixa arborização/NDVI indisponíveis.",
        ),
        (
            "Tendências, sazonalidade, períodos de maior risco e previsão climática/sazonal",
            "Parcial",
            "Período crítico jul–nov; El Niño 2026–2027; predição operacional ~7 dias; chuva 01/09 reduziu temporariamente Tmáx≥37.",
            "Sem outlook sazonal próprio de 5 anos neste arquivo.",
        ),
        (
            "Atendimentos, internações, remoções e óbitos relacionados ao calor (E86/E87/T67/X30 e sensíveis)",
            "Parcial",
            f"Óbitos sensíveis a calor (SIM agregado): {resumo.get('sim_obitos_total')} ({resumo.get('sim_periodo')}). Ver abas 05 e 13.",
            "Remoções e atendimentos ambulatoriais específicos não rotinizados; tipificação calor direto frágil; janela <5 anos completa.",
        ),
        (
            "Distribuição por faixa etária, sexo, município/região, local e grupos vulneráveis",
            "Parcial",
            "Disponível por município/região e proxies demográficos (idosos, crianças, rural, territórios). Ver aba 06.",
            "Sexo/idade/local de ocorrência não consolidados em tabela STAR rotineira neste material.",
        ),
        (
            "Aumento de demanda ou sobrecarga dos serviços em calor extremo",
            "Disponível",
            f"Ocupação IndicaSUS: {round(float((resumo.get('ocup_estado') or {}).get('ocupacao_pct', 0)),1)}% "
            f"({(resumo.get('ocup_estado') or {}).get('leitos_ocupados')}/{(resumo.get('ocup_estado') or {}).get('leitos_existentes')}); "
            f"{resumo.get('n_com_ocupacao')} municípios com ocupação. Ver aba 07.",
            "Cobertura IndicaSUS parcial; pressão SISREG ≠ ocupação.",
        ),
        (
            "Outras análises pertinentes",
            "Disponível",
            f"PM2,5≥25: {resumo.get('n_pm25_ge25')}/142; focos/queimadas; impacto da chuva 01/09. Ver aba 08.",
            "",
        ),
    ]
    _write_kv_table(wsA, "Bloco A — Análise epidemiológica e climática (últimos 5 anos, no mínimo)", rows_a)

    # -------- Limiares --------
    wsL = wb.create_sheet("02_A_Limiares")
    lim = [
        ("Critério", "Parâmetro", "Valor / definição", "Área abrangida"),
        ("UTCI/proxy", "Verde", "≤ 26 °C", "Estado / município"),
        ("UTCI/proxy", "Amarela", "≤ 32 °C", "Estado / município"),
        ("UTCI/proxy", "Laranja", "≤ 38 °C", "Estado / município"),
        ("UTCI/proxy", "Vermelha", "≤ 46 °C", "Estado / município"),
        ("UTCI/proxy", "Roxa", "> 46 °C ou combinação de pilares", "Estado / município"),
        ("Tmáx fallback", "Amarela", "≥ 37 °C", "Estado / município"),
        ("Tmáx fallback", "Laranja", "≥ 39 °C", "Estado / município"),
        ("Tmáx fallback", "Vermelha", "≥ 41 °C", "Estado / município"),
        ("Tmáx fallback", "Roxa", "≥ 43 °C", "Estado / município"),
        ("Risco cumulativo 3d", "Umbral temperatura", "39 °C", "Estado / município"),
        ("Risco cumulativo 3d", "Amarela / Laranja / Vermelha / Roxa", "≥3 / ≥7 / ≥12 / ≥18", "Estado / município"),
        ("EHF adaptado", "Positivo", "> 0", "Estado / município"),
        ("EHF adaptado", "Persistência emergência", "5 dias", "Estado / município"),
        ("Onda de calor", "Duração mínima", "≥ 2 dias consecutivos acima do P95 de temperatura média", "Município"),
        ("Onda de calor", "Intensidade / severidade", "Indicadores intensidade_onda_calor e severidade_onda_calor", "Município"),
        ("Sazonalidade", "Período crítico", "Julho a novembro", "Estado"),
    ]
    _write_kv_table(wsL, "Definição dos limiares utilizados para caracterizar onda de calor", lim)

    # -------- Municipal exposição --------
    wsM = wb.create_sheet("03_A_Municipal_Exposicao")
    cols_pref = [
        c
        for c in [
            "cod_ibge",
            "municipio",
            "regional_saude",
            "classe_atual",
            "classe_projetada_7d",
            "tmax",
            "utci_proxy",
            "umidade_media",
            "risco_cumulativo_3d",
            "onda_calor_flag",
            "duracao_onda_calor_dias",
            "pm25_ugm3",
            "focos_queimadas_7d",
            "densidade",
            "idosos_pct",
            "criancas_0_4_pct",
            "rural_pct",
            "indice_vulnerabilidade_calor",
            "populacao",
            "n_territorios_tradicionais",
            "ocupacao_leitos_pct",
            "pressao_calor_pct",
            "indice_prioridade",
        ]
        if c in mun.columns
    ]
    _write_df(wsM, mun[cols_pref] if cols_pref else mun, freeze="D2")
    wsM.cell(1, len(cols_pref) + 2 if cols_pref else 2, "NOTA: sem camada de ilha de calor / arborização neste arquivo.")

    # -------- Tendências --------
    wsT = wb.create_sheet("04_A_Tendencias")
    tend = [
        ("Indicador", "Valor"),
        ("Período crítico operacional", "Julho a novembro"),
        ("Contexto climático", "El Niño 2026–2027"),
        ("Classe vermelho/roxo atual", f"{resumo.get('vr_atual')}/142"),
        ("Classe vermelho/roxo projeção ~7d", f"{resumo.get('vr_proj')}/142"),
        ("Distribuição atual", str(resumo.get("dist_atual"))),
        ("Previsão operacional disponível", "Sim (~7 dias por município)"),
        ("Previsão climática sazonal oficial embutida", "Não (apenas referência qualitativa CPTEC/INMET/FUNCEME)"),
        ("Série histórica ≥5 anos", "Não disponível neste material"),
    ]
    _write_kv_table(wsT, "Tendências, sazonalidade, períodos de maior risco e previsão", tend)
    chuva = pd.DataFrame(resumo.get("chuva_recente") or [])
    if not chuva.empty:
        start = 14
        wsT.cell(start, 1, "Evidência recente — chuva e Tmáx (impacto 01/09)").font = SECTION_FONT
        for r_idx, row in enumerate(dataframe_to_rows(chuva, index=False, header=True), start + 2):
            for c_idx, val in enumerate(row, 1):
                cell = wsT.cell(r_idx, c_idx, None if pd.isna(val) else val)
                cell.border = THIN
                if r_idx == start + 2:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

    # -------- Saúde calor --------
    wsS = wb.create_sheet("05_A_Saude_Calor")
    saude_rows = [
        ("Indicador", "Valor", "Observação"),
        ("CIDs monitorados (desidratação/calor)", "E86, E87, T67, X30", "Catálogo operacional"),
        ("Óbitos sensíveis a calor (SIM agregado)", resumo.get("sim_obitos_total"), str(resumo.get("sim_periodo"))),
        ("Óbitos tipados como calor direto (agregado)", resumo.get("sim_obitos_calor_susp"), "Tipificação frágil no agregado atual"),
        ("Atendimentos ambulatoriais específicos por calor", "Indisponível", "Sem série estadual rotineira neste material"),
        ("Internações específicas por calor", "Parcial", "Depende DW SIH; não consolidado 5 anos neste arquivo"),
        ("Remoções", "Indisponível", "Sem padronização para o tema"),
    ]
    grupos = resumo.get("saude_calor_grupos") or {}
    for k, v in grupos.items():
        saude_rows.append((f"Grupo saúde–calor: {k}", v, "Série mensal estadual consolidada"))
    _write_kv_table(wsS, "Dados de atendimentos, internações, remoções e óbitos possivelmente relacionados ao calor", saude_rows)

    # -------- Distribuição municipal --------
    wsD = wb.create_sheet("06_A_Distribuicao_Mun")
    dist_cols = [
        c
        for c in [
            "cod_ibge",
            "municipio",
            "regional_saude",
            "classe_atual",
            "tmax",
            "idosos_pct",
            "criancas_0_4_pct",
            "rural_pct",
            "indice_vulnerabilidade_calor",
            "n_territorios_tradicionais",
            "obitos_total_sim",
            "obitos_calor_suspeitos_sim",
            "saude_calor_obitos_sensivel_calor",
        ]
        if c in mun.columns
    ]
    _write_df(wsD, mun[dist_cols] if dist_cols else mun, freeze="D2")
    note_row = (len(mun) + 3) if not mun.empty else 3
    wsD.cell(
        note_row,
        1,
        "NOTA: distribuição por faixa etária, sexo e local de ocorrência não consolidada neste material (existe no SIM bruto).",
    )

    # -------- Sobrecarga --------
    wsO = wb.create_sheet("07_A_Sobrecarga")
    ocup = resumo.get("ocup_estado") or {}
    sob = [
        ("Indicador", "Valor"),
        ("Ocupação estadual IndicaSUS (%)", round(float(ocup.get("ocupacao_pct", 0)), 1) if ocup else "ND"),
        ("Leitos ocupados", ocup.get("leitos_ocupados")),
        ("Leitos existentes", ocup.get("leitos_existentes")),
        ("Municípios com ocupação", ocup.get("municipios_com_ocupacao")),
        ("Fonte", ocup.get("fonte")),
        ("Observação", "Ocupação hospitalar (IndicaSUS) ≠ pressão hospitalar (SISREG)"),
    ]
    _write_kv_table(wsO, "Sobrecarga / demanda dos serviços de saúde", sob)
    sob_cols = [
        c
        for c in [
            "cod_ibge",
            "municipio",
            "regional_saude",
            "classe_atual",
            "ocupacao_leitos_pct",
            "fonte_ocupacao",
            "pressao_calor_pct",
            "semaforo_pressao",
            "indice_pressao_saude",
            "tmax",
        ]
        if c in mun.columns
    ]
    if sob_cols:
        start = 12
        wsO.cell(start, 1, "Detalhamento municipal").font = SECTION_FONT
        det = mun[sob_cols]
        for r_idx, row in enumerate(dataframe_to_rows(det, index=False, header=True), start + 2):
            for c_idx, val in enumerate(row, 1):
                cell = wsO.cell(r_idx, c_idx, None if pd.isna(val) else val)
                cell.border = THIN
                if r_idx == start + 2:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                elif r_idx % 2 == 0:
                    cell.fill = ALT

    # -------- Outras análises --------
    wsX = wb.create_sheet("08_A_Outras_Analises")
    outras = [
        ("Análise", "Resultado"),
        ("Mun. PM2,5 ≥ 25 µg/m³", resumo.get("n_pm25_ge25")),
        ("Mun. Tmáx ≥ 37 °C", resumo.get("n_tmax_ge37")),
        ("Flag onda P95≥2d (rodada)", resumo.get("n_onda_flag")),
        ("Cuiabá pico 31/08", "40,4 °C"),
        ("Impacto chuva 01/09", "Redução temporária de municípios com Tmáx≥37 e da intensidade térmica estadual"),
        ("Qualidade do ar + calor", "Risco multifatorial (calor + fumaça/queimadas)"),
    ]
    _write_kv_table(wsX, "Outras análises pertinentes", outras)
    top_pm = pd.DataFrame(resumo.get("top_pm25") or [])
    if not top_pm.empty:
        wsX.cell(12, 1, "Top 10 PM2,5").font = SECTION_FONT
        for r_idx, row in enumerate(dataframe_to_rows(top_pm, index=False, header=True), 14):
            for c_idx, val in enumerate(row, 1):
                cell = wsX.cell(r_idx, c_idx, None if pd.isna(val) else val)
                cell.border = THIN
                if r_idx == 14:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

    # -------- Bloco B --------
    wsB = wb.create_sheet("09_B_Caracterizacao")
    rows_b = [
        ("Item do Anexo", "Status", "Dados / resposta", "Lacuna"),
        (
            "Principais impactos diretos à saúde relacionados ao calor já observados",
            "Parcial",
            f"Monitoramento de desidratação/insolação/exaustão e agravamentos cardio/respiratório/renal. "
            f"SIM óbitos sensíveis: {resumo.get('sim_obitos_total')} ({resumo.get('sim_periodo')}). "
            f"Exposição SE 35: {resumo.get('vr_atual')}/142 em vermelho/roxo.",
            "Atribuição causal exclusiva ao calor exige estudo adicional; tipificação T67/X30 frágil.",
        ),
        (
            "Possíveis impactos indiretos (hídrica, água/alimentos, DTHA, qualidade do ar, serviços essenciais)",
            "Parcial",
            f"Qualidade do ar: {resumo.get('n_pm25_ge25')}/142 com PM2,5≥25. Hidro/estiagem parcial. DTHA parcial.",
            "Qualidade da água/alimentos e interrupção de serviços essenciais sem série consolidada neste material.",
        ),
        (
            "Fatores que podem ampliar o risco",
            "Parcial",
            "Baixa umidade (INMET); estiagem; trabalho ao ar livre; densidade urbana; calor+fumaça.",
            "Moradia sem ventilação, eventos de massa, abrigos/alojamentos sem indicador municipal rotineiro.",
        ),
        (
            "Grupos com maior vulnerabilidade",
            "Parcial",
            "Ver aba 10_B_Grupos_Vulneraveis e dados municipais de idosos/crianças/rural/territórios.",
            "Gestantes, rua, PPL, imunossuprimidos sem dado rotineiro.",
        ),
        (
            "Outras características relevantes do território",
            "Disponível",
            "El Niño; emergência ambiental por queimadas; arco leste/nordeste com maior Tmáx recente; Baixada Cuiabana em risco persistente.",
            "",
        ),
    ]
    _write_kv_table(wsB, "Bloco B — Caracterização do evento no território e seus impactos à saúde", rows_b)

    # -------- Grupos vulneráveis --------
    wsG = wb.create_sheet("10_B_Grupos_Vulneraveis")
    grupos_v = [
        ("Grupo (Anexo)", "Dado disponível?", "Como está representado neste arquivo"),
        ("Crianças", "Sim", "criancas_0_4_pct na aba municipal"),
        ("Idosos", "Sim", "idosos_pct / indice_vulnerabilidade_calor"),
        ("Gestantes", "Não", "Lacuna"),
        ("Pessoas com doenças crônicas", "Não", "Lacuna (dado individual)"),
        ("Imunossuprimidos", "Não", "Lacuna"),
        ("Pessoas em situação de rua", "Não", "Lacuna"),
        ("População privada de liberdade", "Não", "Lacuna"),
        ("População indígena", "Parcial", "n_territorios_tradicionais / mapas boletim"),
        ("Trabalhadores expostos ao calor", "Parcial", "Orientação Saúde do Trabalhador / AdaptaSUS"),
        ("Pessoas em moradias precárias", "Parcial", "Proxy ruralidade + índice de vulnerabilidade"),
    ]
    _write_kv_table(wsG, "Grupos com maior vulnerabilidade no estado", grupos_v)
    topv = pd.DataFrame(resumo.get("top_vuln") or [])
    if not topv.empty:
        wsG.cell(16, 1, "Top 10 índice de vulnerabilidade ao calor").font = SECTION_FONT
        for r_idx, row in enumerate(dataframe_to_rows(topv, index=False, header=True), 18):
            for c_idx, val in enumerate(row, 1):
                cell = wsG.cell(r_idx, c_idx, None if pd.isna(val) else val)
                cell.border = THIN
                if r_idx == 18:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

    # -------- Bloco C --------
    wsC = wb.create_sheet("11_C_Vigilancia_Resposta")
    rows_c = [
        ("Item do Anexo", "Status", "Dados / resposta", "Lacuna"),
        (
            "Estrutura atual da vigilância (federal, estadual e municipal)",
            "Disponível",
            "Federal: INMET, MS/AdaptaSUS, Rede CIEVS. Estadual: CIEVS-MT, Sala de Situação (Portaria 0590/2026), ARARAS MT, Vigidesastres, áreas SES. Municipal: VS, Defesa Civil, APS/rede.",
            "Inventário formal de capacidades municipais incompleto.",
        ),
        (
            "Disponibilidade de insumos, exames laboratoriais e protocolos",
            "Parcial",
            "Protocolos/limiares operacionais disponíveis (aba 02). Módulos de estoque/autonomia parciais no Plano El Niño.",
            "Inventário quantitativo completo por município específico para ondas de calor indisponível.",
        ),
        (
            "Recursos humanos envolvidos e lacunas",
            "Lacuna",
            "Envolvimento CIEVS-MT, Sala, vigilância/assistência SES e focais municipais/regionais.",
            "Sem quadro RH dedicado consolidado neste material.",
        ),
        (
            "Articulação com outros setores (meio ambiente, educação, assistência social, etc.)",
            "Disponível",
            "SEMA, CBM, Educação, Assistência Social, Defesa Civil, Saúde do Trabalhador, DSEI/SESAI, áreas SES do Plano El Niño.",
            "",
        ),
        (
            "Planos de contingência, políticas, normas e legislação",
            "Disponível",
            "Portaria 0590/2026/GBSES; Plano El Niño SES-MT; Decreto 2.015/2026; AdaptaSUS; ARARAS MT.",
            "Planos municipais de calor extremo não catalogados em base única.",
        ),
        (
            "Principais desafios e oportunidades de melhoria",
            "Disponível",
            "Desafios: série ≥5 anos; mapas ilha de calor/arborização; tipificação idade/sexo; inventário RH/insumos/planos municipais. "
            "Oportunidades: rotina semanal Sala/ARARAS; backfill 5 anos; priorização exposição×vulnerabilidade.",
            "",
        ),
    ]
    _write_kv_table(wsC, "Bloco C — Sistema de Vigilância e Capacidade de Resposta", rows_c)

    # -------- Alertas INMET --------
    wsI = wb.create_sheet("12_Alertas_INMET")
    ia = read_table("inmet_alertas")
    if ia is not None and not ia.empty:
        cols = [
            c
            for c in [
                "data_emissao",
                "inicio_vigencia",
                "fim_vigencia",
                "evento",
                "nivel_alerta",
                "severidade",
                "area_mt",
                "status",
                "fonte",
                "link",
            ]
            if c in ia.columns
        ]
        _write_df(wsI, ia[cols].copy(), freeze="A2")
    else:
        wsI["A1"] = "Sem alertas INMET disponíveis na base."

    # -------- Série saúde mensal --------
    wsSM = wb.create_sheet("13_Serie_Saude_Mensal")
    se = read_table("saude_calor_serie_estado")
    if se is not None and not se.empty:
        _write_df(wsSM, se.sort_values(["mes", "grupo_agravo_calor"]), freeze="A2")
    else:
        wsSM["A1"] = "Série mensal saúde–calor indisponível."

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)
    print("sheets=", wb.sheetnames)


if __name__ == "__main__":
    main()
