"""Gera planilha XLSX do levantamento STAR Ondas de Calor."""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "output" / "star" / "STAR_ondas_calor_municipal_SE35_2026.csv"
JSON_PATH = ROOT / "data" / "output" / "star" / "STAR_resumo_indicadores.json"
OUT_PATH = ROOT / "data" / "output" / "star" / "STAR_Ondas_de_Calor_MT_SE35_2026.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1D357F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="F7F9FC")
THIN = Border(
    left=Side(style="thin", color="C8D2E6"),
    right=Side(style="thin", color="C8D2E6"),
    top=Side(style="thin", color="C8D2E6"),
    bottom=Side(style="thin", color="C8D2E6"),
)
STATUS_FILLS = {
    "Disponível": PatternFill("solid", fgColor="C6EFCE"),
    "Disponível (proxy)": PatternFill("solid", fgColor="C6EFCE"),
    "Parcial": PatternFill("solid", fgColor="FFEB9C"),
    "Lacuna": PatternFill("solid", fgColor="FFC7CE"),
}


def _write_df(ws, df: pd.DataFrame, freeze: str | None = "A2") -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, None if pd.isna(val) else val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif r_idx % 2 == 0:
                cell.fill = ALT_FILL
    if len(df.columns):
        ws.auto_filter.ref = ws.dimensions
    if freeze:
        ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 30
    for i, col in enumerate(df.columns, 1):
        sample = [str(col)] + [str(x) for x in df[col].head(25).tolist()]
        width = min(42, max(10, max(len(s) for s in sample) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def main() -> None:
    if not CSV_PATH.exists():
        raise SystemExit(f"CSV não encontrado: {CSV_PATH}. Rode scripts/exportar_star_ondas_calor.py")

    df = pd.read_csv(CSV_PATH)
    resumo = json.loads(JSON_PATH.read_text(encoding="utf-8")) if JSON_PATH.exists() else {}

    wb = Workbook()

    ws = wb.active
    ws.title = "Municipal"
    _write_df(ws, df, freeze="E2")

    ws2 = wb.create_sheet("Resumo_Estadual")
    ws2["A1"] = "Levantamento STAR — Ondas de Calor — SE 35/2026"
    ws2["A1"].font = Font(bold=True, size=14, color="1D357F")
    ws2.merge_cells("A1:B1")
    ws2["A2"] = "Fonte: ARARAS MT / CIEVS-MT"
    ws2["A2"].font = Font(italic=True, size=10)

    rows_resumo = [
        ("Indicador", "Valor"),
        ("Municípios (n)", resumo.get("n_mun")),
        ("Vermelho/roxo atual", f"{resumo.get('vr_atual')}/142"),
        ("Vermelho/roxo projeção ~7d", f"{resumo.get('vr_proj')}/142"),
        ("Distribuição atual", str(resumo.get("dist_atual"))),
        ("Tmáx máxima (°C)", resumo.get("tmax_max")),
        ("Mun. Tmáx ≥ 37 °C", resumo.get("n_tmax_ge37")),
        ("Mun. PM2,5 ≥ 25 µg/m³", resumo.get("n_pm25_ge25")),
        ("Flag onda P95≥2d", resumo.get("n_onda_flag")),
        ("Mun. com ocupação IndicaSUS", resumo.get("n_com_ocupacao")),
    ]
    ocup = resumo.get("ocup_estado") or {}
    if ocup:
        rows_resumo.extend(
            [
                ("Ocupação estadual (%)", round(float(ocup.get("ocupacao_pct", 0)), 1)),
                (
                    "Leitos ocupados / existentes",
                    f"{ocup.get('leitos_ocupados')}/{ocup.get('leitos_existentes')}",
                ),
                ("Fonte ocupação", ocup.get("fonte")),
            ]
        )
    hist = resumo.get("hist_clima_periodo") or []
    if len(hist) == 2:
        rows_resumo.append(("Histórico clima (período)", f"{hist[0]} a {hist[1]}"))
    sim_p = resumo.get("sim_periodo") or []
    if len(sim_p) == 2:
        rows_resumo.append(("SIM óbitos sensíveis (período)", f"{sim_p[0]} a {sim_p[1]}"))
        rows_resumo.append(("SIM óbitos total (extração)", resumo.get("sim_obitos_total")))

    for r_idx, row in enumerate(rows_resumo, 4):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(r_idx, c_idx, val)
            cell.border = THIN
            if r_idx == 4:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif c_idx == 1:
                cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 58

    ws3 = wb.create_sheet("Top10_Tmax")
    _write_df(ws3, pd.DataFrame(resumo.get("top_tmax") or []))

    ws4 = wb.create_sheet("Top10_Vulnerabilidade")
    _write_df(ws4, pd.DataFrame(resumo.get("top_vuln") or []))

    ws5 = wb.create_sheet("Top10_PM25")
    _write_df(ws5, pd.DataFrame(resumo.get("top_pm25") or []))

    ws6 = wb.create_sheet("Checklist_STAR")
    ws6["A1"] = "Checklist Anexo STAR — status"
    ws6["A1"].font = Font(bold=True, size=12, color="1D357F")
    check_rows = [
        ("Bloco", "Item", "Status"),
        ("A", "Série ondas/Tmáx/anomalias/alertas ≥5 anos", "Parcial"),
        ("A", "Limiares duração/intensidade/área", "Disponível"),
        ("A", "Mapas exposição/vulnerabilidade (ilha calor/arborização)", "Parcial"),
        ("A", "Tendências/sazonalidade/previsão", "Parcial"),
        ("A", "Atendimentos/internações/remoções/óbitos calor", "Parcial"),
        ("A", "Distribuição idade/sexo/mun/regional", "Parcial"),
        ("A", "Sobrecarga assistencial", "Disponível (proxy)"),
        ("B", "Impactos diretos à saúde", "Parcial"),
        ("B", "Impactos indiretos (água/ar/DTHA)", "Parcial"),
        ("B", "Amplificadores de risco", "Parcial"),
        ("B", "Grupos vulneráveis", "Parcial"),
        ("C", "Estrutura vigilância federal/estadual/municipal", "Disponível"),
        ("C", "Insumos/exames/protocolos", "Parcial"),
        ("C", "Recursos humanos", "Lacuna"),
        ("C", "Articulação intersetorial", "Disponível"),
        ("C", "Planos/normas/legislação", "Disponível"),
        ("C", "Desafios e oportunidades", "Disponível"),
    ]
    for r_idx, row in enumerate(check_rows, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws6.cell(r_idx, c_idx, val)
            cell.border = THIN
            if r_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif c_idx == 3 and val in STATUS_FILLS:
                cell.fill = STATUS_FILLS[val]
    ws6.column_dimensions["A"].width = 10
    ws6.column_dimensions["B"].width = 62
    ws6.column_dimensions["C"].width = 20
    ws6["A22"] = "Nota completa: docs/apresentacoes/STAR_Ondas_de_Calor_MT_levantamento.md"
    ws6["A22"].font = Font(italic=True, size=9)

    ws7 = wb.create_sheet("Chuva_recente")
    _write_df(ws7, pd.DataFrame(resumo.get("chuva_recente") or []))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(OUT_PATH)
    print("sheets=", wb.sheetnames)
    print("municipal_rows=", len(df))


if __name__ == "__main__":
    main()
